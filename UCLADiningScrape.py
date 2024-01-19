#!/usr/bin/env python
# coding: utf-8

# In[12]:


import lxml
import openpyxl
import pandas as pd
import ssl
import datetime as dt
import pytz
import os

#os.chdir(r"C:\Users\begole\Desktop\Dining Hall Scraper")


# In[13]:


ssl._create_default_https_context = ssl._create_unverified_context

scraper = pd.read_html("https://menu.dining.ucla.edu/Hours")


# In[14]:

print("Running dining hall checks...")

#for idx, table in enumerate (scraper):
 #   print("-----------------------------------------------------")
  #  print(idx)
   # print(table)
    


# In[15]:


#get the current time to find out the meal period and to store in the correct part of the xlsx
curTime = dt.datetime.now(pytz.timezone('America/Los_Angeles'))
#print(curTime)

#Store a variable describing what mealtime it is
mealtime = 'Unknown'

#Figure out what meal period it is
#error case if somehow not in daytime
if curTime.hour <0  or curTime.hour > 24:
    print('ERROR IN HOUR')
    
elif curTime.hour >= 7 and curTime.hour <= 10 :
    #print('Breakfast')
    curHours = scraper[0][['Breakfast']].copy()
    mealtime = 'Breakfast'
    
elif curTime.hour >= 11 and curTime.hour <= 16:
    #print('Lunch/Brunch')
    curHours = scraper[0][['Lunch/Brunch']].copy()
    mealtime = 'Lunch/Brunch'

elif curTime.hour >= 17 and curTime.hour <21:
    #print('Dinner')
    curHours = scraper[0][['Dinner']].copy()
    mealtime = 'Dinner'

elif curTime.hour >= 21 and curTime.hour <= 24:
    #print('Extended Dinner')
    curHours = scraper[0][['Extended Dinner']].copy()
    mealtime = 'Extended Dinner'


else:
    print('CLOSED')
        


# In[16]:


if curHours.index.stop > 10:
    curHours.drop([10,11,12], inplace = True)


# In[17]:


for ind in curHours.index:
    if curHours.loc[ind, mealtime].find('%') != -1:
        curHours.loc[ind, mealtime] = curHours.loc[ind, mealtime].split(' ')[-1]
        #print(len(curHours.loc[ind, mealtime]))
        curHours.loc[ind, mealtime] = curHours.loc[ind, mealtime][:len(curHours.loc[ind, mealtime])-1]

    else:
        curHours.loc[ind, mealtime] = ''
    #curHours.loc[ind, mealtime]=''.join(i for i in curHours.loc[ind, mealtime] if i.isdigit())

                                            
curHours = curHours.T
#curHours


# In[18]:


weekday = "Error"
match curTime.weekday():
    case 0:
        weekday = "Monday"
    case 1:
        weekday = "Tuesday"
    case 2:
        weekday = "Wednesday"
    case 3:
        weekday = "Thursday"
    case 4:
        weekday = "Friday"
    case 5:
        weekday = "Saturday"
    case 6:
        weekday = "Sunday"
    case _: 
        print("Error in Weekday")

wb = openpyxl.load_workbook(filename="DiningHallHourData.xlsx")
wb.active = wb[weekday]

sheet = wb.active


# In[19]:


rIdx = 0

if curTime.hour < 7:
    print("Error in Hour")
    exit()
else:
    rIdx = 4*(curTime.hour-7)
    if curTime.minute < 15:
        rIdx += 2
    elif curTime.minute < 30:
        rIdx += 3
    elif curTime.minute < 45:
        rIdx += 4
    elif curTime.minute < 60:
        rIdx += 5

#print(rIdx)
    


# In[20]:

'''
for col in range (14, 24):
    curcell = sheet.cell(row = rIdx, column = col)
    if curcell.value == None:
        curcell.value = 0
    if curHours.loc[mealtime, col-2].isdigit() == True:
        curcell.value += 1
    #print(curcell.value)
'''


# In[21]:


for col in range (2, 12):

    curcell = sheet.cell(row = rIdx, column = col)
    numElem = sheet.cell(row = rIdx, column = col+12)
    if curcell.value == 0:
        if curHours.loc[mealtime, col-2] != '':
            curcell.value += int(curHours.loc[mealtime, col-2])
            numElem.value += 1
            
    elif  curHours.loc[mealtime, col-2].isdigit() == True:
        curcell.value = int(curcell.value) + (int(curHours.loc[mealtime, col-2])-int(curcell.value))/int(sheet.cell(row = rIdx, column = col+12).value)
#        print("Error in retrieved value")
        #curcell.value = float(curcell.value)
        numElem.value += 1
        sheet.cell(row = rIdx, column = col).value = float(curcell.value)
    else:
        break
    
    print(curcell.value)


#NMstartIdx = "N" + str(rIdx)
#NMstopIdx = "W" + str(rIdx)

#sheet[NMstartIDX,":",NMstopIdx]



# In[22]:


wb.save("DiningHallHourData.xlsx")

print("Finished checking.")

# In[ ]:




